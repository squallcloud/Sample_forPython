import openpyxl

xl_path = "sample.xlsx"

wb = openpyxl.load_workbook(xl_path)
for sheet_name in wb.sheetnames:
  print(sheet_name)
  sheet = wb[sheet_name]
  print(sheet.max_column)
  print(sheet.max_row)

  for row_idx in range(sheet.max_row):
    for column_idx in range(sheet.max_column):
      cell = sheet.cell(row=row_idx+1, column=column_idx+1)
      value = cell.value
      if type(value) is str:
        repalce_value = value.replace("hoge", "foo")
        cell.value = repalce_value
        cell.value = 10

wb.save(xl_path)
wb.close()